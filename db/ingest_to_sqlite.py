import os
import re
import sqlite3
import pandas as pd
import openpyxl

def clean_course_code(raw_code):
    """
    Normalizes course code by extracting the main prefix and number.
    E.g., 'MA 104 (E)' or 'MA 104 (H)' becomes 'MA 104'.
    """
    if pd.isna(raw_code) or not isinstance(raw_code, str):
        return None
    m = re.search(r'([A-Za-z]{2,3})\s*(\d{3})', raw_code)
    if m:
        return f"{m.group(1).upper()} {m.group(2)}"
    return raw_code.strip().upper()

def extract_slots(text):
    """
    Extracts comma-separated timetable slot codes from schedule strings.
    E.g., 'P1,P2\n(Central Arcade)' becomes 'P1,P2'.
    """
    if pd.isna(text) or not isinstance(text, str):
        return None
    text = text.strip()
    # Match slot patterns like A1, A2, B1, etc. at the start of the cell
    m = re.match(r'^([A-P][1-2](?:\s*,\s*[A-P][1-2])*)', text, re.IGNORECASE)
    if m:
        return ",".join([s.strip().upper() for s in m.group(1).split(",")])
    return None

def ingest_excel(file_path, semester, conn):
    print(f"Processing sheet: {file_path} ({semester} semester)...")
    
    # Load with openpyxl to get hyperlinks
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb["Time table"]
    
    # Find column indices (1-based for openpyxl)
    header_row = [cell.value for cell in sheet[1]]
    
    col_mapping = {
        "course_number": "Course Number",
        "course_name": "Course Name",
        "L": "L",
        "T": "T",
        "P": "P",
        "C": "C",
        "instructor": "Name of the Instructors and Tutors",
        "lecture": "Lecture",
        "tutorial": "Tutorial",
        "lab": "Lab",
        "link_to_plan": "Link To Course Plan",
        "hss_bs": "HSS/BS elective"
    }
    
    col_indices = {}
    for col_key, col_name in col_mapping.items():
        try:
            col_indices[col_key] = header_row.index(col_name) + 1
        except ValueError:
            col_indices[col_key] = None
            
    print("Column indices extracted:", col_indices)
    
    course_code_regex = r'[A-Za-z]{2,3}\s*\d{3}'
    cursor = conn.cursor()
    
    rows_added = 0
    # Iterate from row 2 (skipping header)
    for r in range(2, sheet.max_row + 1):
        course_num_cell = sheet.cell(row=r, column=col_indices["course_number"])
        course_num = course_num_cell.value
        
        # Skip if course number is null or doesn't look like a course code
        if not course_num or not re.search(course_code_regex, str(course_num)):
            continue
            
        course_num = str(course_num).strip()
        clean_code = clean_course_code(course_num)
        
        # Calculate semester number for BTech courses (course numbers < 500)
        m_num = re.search(r'\d{3}', clean_code)
        semester_num = None
        if m_num:
            code_digits = int(m_num.group())
            # Calculate BTech Core Semester Number: 100-lvl Odd=1, Even=2; 200-lvl Odd=3, Even=4; etc.
            year = code_digits // 100
            if year in [1, 2, 3, 4]:
                sem_offset = 0 if semester.lower() == "odd" else 1
                semester_num = (year - 1) * 2 + 1 + sem_offset
        
        # Course Name
        course_name = sheet.cell(row=r, column=col_indices["course_name"]).value
        if course_name:
            course_name = str(course_name).strip()
            
        # L, T, P, C (handling numeric types safely)
        l_val = sheet.cell(row=r, column=col_indices["L"]).value
        t_val = sheet.cell(row=r, column=col_indices["T"]).value
        p_val = sheet.cell(row=r, column=col_indices["P"]).value
        c_val = sheet.cell(row=r, column=col_indices["C"]).value
        
        l_val = int(l_val) if l_val is not None and str(l_val).replace('.','').isdigit() else 0
        t_val = int(t_val) if t_val is not None and str(t_val).replace('.','').isdigit() else 0
        p_val = int(p_val) if p_val is not None and str(p_val).replace('.','').isdigit() else 0
        c_val = int(c_val) if c_val is not None and str(c_val).replace('.','').isdigit() else 0
        
        # Instructor
        instructor = sheet.cell(row=r, column=col_indices["instructor"]).value
        if instructor:
            instructor = str(instructor).strip()
            
        # Slots
        lec_val = sheet.cell(row=r, column=col_indices["lecture"]).value
        tut_val = sheet.cell(row=r, column=col_indices["tutorial"]).value
        lab_val = sheet.cell(row=r, column=col_indices["lab"]).value
        
        lecture_slots = extract_slots(lec_val)
        tutorial_slots = extract_slots(tut_val)
        lab_slots = extract_slots(lab_val)
        
        # Hyperlinks extraction
        link_cell = sheet.cell(row=r, column=col_indices["link_to_plan"])
        link_url = None
        if link_cell.hyperlink:
            link_url = link_cell.hyperlink.target
        elif link_cell.value and str(link_cell.value).startswith("http"):
            link_url = str(link_cell.value).strip()
            
        # HSS/BS Elective status
        hss_bs_val = None
        if col_indices["hss_bs"]:
            hss_bs_val = sheet.cell(row=r, column=col_indices["hss_bs"]).value
            if hss_bs_val:
                hss_bs_val = str(hss_bs_val).strip()
                
        # Insert into Database
        cursor.execute("""
            INSERT INTO courses (
                course_code, clean_course_code, title, instructor, semester, semester_num,
                L, T, P, C, course_plan_url, lecture_slots, tutorial_slots, lab_slots, hss_bs_elective
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            course_num, clean_code, course_name, instructor, semester, semester_num,
            l_val, t_val, p_val, c_val, link_url, lecture_slots, tutorial_slots, lab_slots, hss_bs_val
        ))
        rows_added += 1
        
    conn.commit()
    print(f"Successfully loaded {rows_added} course rows from {file_path}.")

def main():
    db_file = "acadagent.db"
    
    # Remove existing db if we want a fresh start
    if os.path.exists(db_file):
        os.remove(db_file)
        print("Removed existing acadagent.db database.")
        
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    
    # Create tables
    cursor.execute("""
        CREATE TABLE courses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            course_code TEXT,
            clean_course_code TEXT,
            title TEXT,
            instructor TEXT,
            semester TEXT,
            semester_num INTEGER,
            L INTEGER,
            T INTEGER,
            P INTEGER,
            C INTEGER,
            course_plan_url TEXT,
            lecture_slots TEXT,
            tutorial_slots TEXT,
            lab_slots TEXT,
            hss_bs_elective TEXT
        )
    """)
    conn.commit()
    print("Created table 'courses' in SQLite.")
    
    # Ingest even and odd semesters
    ingest_excel("sem_odd.xlsx", "odd", conn)
    ingest_excel("sem_even.xlsx", "even", conn)
    
    # Verification query
    print("\n--- Verifying SQLite Insertion ---")
    cursor.execute("SELECT COUNT(*) FROM courses;")
    total_courses = cursor.fetchone()[0]
    print(f"Total course records in database: {total_courses}")
    
    print("\nSample records:")
    df_verify = pd.read_sql_query("SELECT course_code, clean_course_code, title, semester, C, course_plan_url, lecture_slots FROM courses LIMIT 5;", conn)
    print(df_verify.to_string())
    
    conn.close()

if __name__ == "__main__":
    main()
